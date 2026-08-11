"""A standard that ingested nothing must not report success.

Before these guards, `StandardsLoader.ingest_standard` answered a zero-chunk parse by inventing
one chunk containing only the title the uploader typed, saving it, and returning 200. A scanned
PDF — the single most likely bad input for engineering standards — therefore produced a standard
that appeared in the list, reported a chunk, and held none of its own content. Permanently:
re-uploading the same file hits the duplicate-hash bypass and never re-parses.

That is the same failure this project already paid for with SHA-256 embeddings: returning
something plausible rather than failing. Nothing downstream could tell the difference.

The parser half is tested here directly because it is pure; the loader half needs Mongo, which
this suite does not have, so its guard is exercised through the parse result it branches on.
"""

import pytest

from services.backend.infrastructure.audit.standards_parser import (
    SUPPORTED_STANDARD_FORMATS,
    StandardIngestError,
    StandardsParser,
)


@pytest.fixture
def sandboxed(tmp_path, monkeypatch):
    """`parse_file` refuses paths outside the storage root; point the root at tmp_path."""
    monkeypatch.setattr(
        "services.backend.infrastructure.audit.standards_parser.validate_sandboxed_path",
        lambda p: p,
    )
    return tmp_path


def test_the_standards_directory_survives_the_sandbox_guard():
    """The loader must write inside the root the guard enforces. It did not.

    `StandardsLoader` built its destination from `settings.STORAGE_ROOT` — which defaults to the
    relative `"./storage"` and so resolves against the backend's working directory — while
    `validate_sandboxed_path` enforces `get_storage_root()`, an absolute path derived from the
    package location. With the backend started from `services/backend/`, every upload died:

        Path Traversal Attempt Blocked: Resolved path
        '...\\services\\backend\\storage\\standards\\<hash>.xls'
        escapes storage root boundary '...\\ai-2d-checker\\storage'

    A security guard reporting a traversal *attempt* for the application's own write path is the
    signature to recognise: it is two components disagreeing about a root, not an attack. The
    same family as [[Gotcha - The Sweep Never Got the Zone Template Seam]] — two callers, one
    concept, no shared definition.
    """
    from services.backend.core.security import validate_sandboxed_path
    from services.backend.infrastructure.audit.standards_loader import standards_storage_dir
    from services.backend.infrastructure.storage.path_resolver import get_storage_root

    target = standards_storage_dir()

    # The guard itself is the assertion; it raises HTTPException(400) on a mismatch.
    validate_sandboxed_path(target / "deadbeef.xlsx")

    assert target.is_relative_to(get_storage_root())


def test_stored_file_path_is_relative_to_the_same_root_readers_use():
    """`StandardDocument.file_path` is stored relative; against the wrong root it resolves nowhere."""
    import os

    from services.backend.infrastructure.audit.standards_loader import standards_storage_dir
    from services.backend.infrastructure.storage.path_resolver import get_storage_root

    dest = standards_storage_dir() / "abc123.xlsx"
    relative = os.path.relpath(dest, get_storage_root())

    assert not os.path.isabs(relative)
    assert (get_storage_root() / relative).resolve() == dest.resolve()


def test_both_excel_formats_are_supported():
    assert "xls" in SUPPORTED_STANDARD_FORMATS
    assert "xlsx" in SUPPORTED_STANDARD_FORMATS


def test_xls_dispatches_to_the_legacy_reader(sandboxed, monkeypatch):
    """`.xls` must reach xlrd, never openpyxl, which cannot read the OLE2 container.

    Asserted at the dispatch rather than by parsing a binary fixture: the routing decision is
    the thing that regressed before (three format lists claimed `.xls` while the only code path
    handed it to openpyxl), and it is checkable without committing a 30 MB customer workbook.
    """
    seen = {}

    def fake_xls(path, meta):
        seen["reader"] = "xls"
        meta["page_count"] = 1
        return iter(())

    monkeypatch.setattr(StandardsParser, "_read_xls_rows", staticmethod(fake_xls))

    f = sandboxed / "KEMCO and JIS Standards.xls"
    f.write_bytes(b"\xd0\xcf\x11\xe0")  # OLE2 magic — a genuine legacy .xls header

    chunks, _ = StandardsParser.parse_file(f)
    assert seen["reader"] == "xls"
    assert chunks == []


def test_xlsx_dispatches_to_the_ooxml_reader(sandboxed, monkeypatch):
    seen = {}

    def fake_xlsx(path, meta):
        seen["reader"] = "xlsx"
        meta["page_count"] = 1
        return iter(())

    monkeypatch.setattr(StandardsParser, "_read_xlsx_rows", staticmethod(fake_xlsx))

    f = sandboxed / "modern.xlsx"
    f.write_bytes(b"PK\x03\x04")

    StandardsParser.parse_file(f)
    assert seen["reader"] == "xlsx"


def test_the_two_readers_share_one_set_of_semantics(sandboxed, monkeypatch):
    """The claim the whole design rests on, asserted rather than trusted.

    `.xls` and `.xlsx` differ only in how a cell is *read*. What a fill colour means, that bold
    becomes markdown, how rows are joined and chunked — all of it is downstream of `_Cell` and
    must therefore be byte-identical for identical input. A second copy of the colour rules
    would be free to disagree, and the disagreement would look like a difference in the
    standards themselves rather than a bug.
    """
    from services.backend.infrastructure.audit.standards_parser import _Cell, _SheetRow

    rows = [
        _SheetRow("Bolting", 1, [_Cell("Chamfer", True, "FFC7CE")]),
        _SheetRow("Bolting", 2, [_Cell("面取り", False, "C6EFCE"), _Cell("C0.5", False, None)]),
    ]

    def reader(path, meta):
        meta["page_count"] = 1
        return iter(rows)

    outputs = []
    for attr, suffix in (("_read_xls_rows", ".xls"), ("_read_xlsx_rows", ".xlsx")):
        monkeypatch.setattr(StandardsParser, attr, staticmethod(reader))
        f = sandboxed / f"same{suffix}"
        f.write_bytes(b"\x00")
        chunks, _ = StandardsParser.parse_file(f)
        outputs.append([c["content"] for c in chunks])

    assert outputs[0] == outputs[1], "the two formats produced different text for one input"
    body = "\n".join(outputs[0])
    assert "[INCORRECT / DANGER FLAG] **Chamfer**" in body
    assert "[CORRECT / STANDARDS COMPLIANT] 面取り | C0.5" in body


def test_unsupported_extension_names_what_is_allowed(sandboxed):
    f = sandboxed / "standard.docx"
    f.write_bytes(b"PK\x03\x04")

    with pytest.raises(StandardIngestError) as err:
        StandardsParser.parse_file(f)
    assert ".pdf" in str(err.value) and ".xlsx" in str(err.value)


def test_a_pdf_with_no_text_layer_yields_no_chunks(sandboxed):
    """The scanned-PDF case. The parser returns nothing; the loader must then refuse.

    A one-page PDF with no text operators stands in for a scan — `extract_text()` returns an
    empty string for both, which is the condition the loader branches on.
    """
    pdf = sandboxed / "scanned.pdf"
    pdf.write_bytes(
        b"%PDF-1.4\n"
        b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
        b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
        b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]>>endobj\n"
        b"trailer<</Root 1 0 R>>\n"
    )

    try:
        chunks, _ = StandardsParser.parse_file(pdf)
    except StandardIngestError:
        # A malformed-enough PDF is rejected outright, which is also a refusal to ingest
        # silently. Either outcome satisfies the property under test.
        return
    assert chunks == [], "a text-free PDF must produce no chunks for the loader to reject"


def test_shift_jis_text_is_decoded_not_mangled(sandboxed):
    """`errors='ignore'` on hard-coded utf-8 was silent corruption in a Japanese CAD shop.

    Mangled text is worse than no text: it is non-empty, so it passes every emptiness check
    downstream and lands a corrupted corpus that looks fine.
    """
    original = "表示外公差は普通公差による\n面取りは糸面取りのこと\n"
    f = sandboxed / "jis.txt"
    f.write_bytes(original.encode("cp932"))

    chunks, meta = StandardsParser.parse_file(f)

    body = "\n".join(c["content"] for c in chunks)
    assert "表示外公差" in body
    assert "糸面取り" in body
    assert meta["encoding"] in ("cp932", "shift_jis")


def test_utf8_still_wins_over_the_fallbacks(sandboxed):
    """The fallback chain must not change how a correct file is read."""
    f = sandboxed / "utf8.md"
    f.write_text("# 規格\n\n面取りのこと\n", encoding="utf-8")

    chunks, meta = StandardsParser.parse_file(f)
    assert meta["encoding"] == "utf-8"
    assert any("面取り" in c["content"] for c in chunks)


def _xlsx_with(rows, images=0):
    import openpyxl

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Rules"
    for row in rows:
        sheet.append(row)
    return wb


def test_excel_reports_images_it_cannot_read(sandboxed):
    """Pictures contribute no text and previously left no trace at all."""
    openpyxl = pytest.importorskip("openpyxl")
    wb = _xlsx_with([["Rule", "Value"], ["Chamfer", "C0.5"]])

    # A 1x1 PNG is enough — the parser counts pictures, it does not decode them.
    png = (
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00"
        b"\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01\x00\x00\x05\x00\x01\r"
        b"\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82"
    )
    img_path = sandboxed / "diagram.png"
    img_path.write_bytes(png)
    wb["Rules"].add_image(openpyxl.drawing.image.Image(str(img_path)), "D2")

    out = sandboxed / "rules.xlsx"
    wb.save(out)

    chunks, meta = StandardsParser.parse_file(out)
    assert meta["images_ignored"] == 1, "a dropped image must be counted, not silently skipped"
    assert any("Chamfer" in c["content"] for c in chunks), "cell text must still be read"


def test_excel_with_no_cell_text_yields_no_chunks(sandboxed):
    """The images-only workbook: nothing readable, so nothing may be reported as ingested."""
    pytest.importorskip("openpyxl")
    wb = _xlsx_with([])
    out = sandboxed / "empty.xlsx"
    wb.save(out)

    chunks, _ = StandardsParser.parse_file(out)
    assert chunks == []


def test_excel_colour_coding_survives_into_the_text(sandboxed):
    """The one genuine advantage Excel has over PDF here; pinned so a refactor cannot drop it."""
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.styles import PatternFill

    wb = _xlsx_with([["Chamfer", "C0.5"]])
    wb["Rules"]["A1"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    out = sandboxed / "coloured.xlsx"
    wb.save(out)

    chunks, _ = StandardsParser.parse_file(out)
    assert "[INCORRECT / DANGER FLAG]" in chunks[0]["content"]
