import re
from collections.abc import Iterator
from pathlib import Path
from typing import Any, NamedTuple

from pypdf import PdfReader

from ...core.security import validate_sandboxed_path
from ...logger import logger


# The single source of truth for what may be ingested. Previously restated in three places —
# this module's dispatch, `StandardsLoader.ingest_standard` and the upload router — and they had
# already drifted: all three advertised `.xls` while no reader existed for it. There is one now,
# so the entry is honest; the constant stays because the drift is what caused the outage.
SUPPORTED_STANDARD_FORMATS = ("pdf", "txt", "md", "xlsx", "xls")


class _Cell(NamedTuple):
    """A cell reduced to the three things chunking needs, independent of which library read it.

    This type is what lets `.xls` and `.xlsx` share their semantics instead of merely resembling
    each other. Both readers produce it; everything that assigns *meaning* consumes it.
    """

    text: str
    bold: bool
    fill_rgb: str | None


class _SheetRow(NamedTuple):
    sheet_name: str
    row_index: int
    cells: list[_Cell]


class StandardIngestError(ValueError):
    """A problem with the uploaded file that the person who uploaded it can act on.

    Separate from every other exception for one reason: the upload router surfaces this message
    verbatim and replaces everything else with an opaque correlation id. That generic
    handler is correct — an unexpected exception can carry a filesystem path or an internal
    detail — but it also meant a wrong-but-fixable file ("this is a scanned PDF", "re-save as
    .xlsx") produced *"Ingestion process failed. Reference: <uuid>"*, which names neither the
    cause nor the fix. Messages raised as this type are written for a human holding the file.
    """


class StandardsParser:
    """
    Safely parses PDF, TXT, Excel and Markdown files inside a sandboxed storage root.
    Extracts structured chunks, metadata, and handles basic format validation.
    """

    @staticmethod
    def parse_file(file_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        """
        Parses a Grounding Standard document based on its extension.
        Returns:
            chunks: List of dictionaries, each containing:
                - content (str)
                - section_header (str or None)
                - metadata (dict with page number, line offsets, etc.)
            metadata: Dict with global properties (author, title, page_count).
        """
        # 1. Enforce sandbox traversal limits
        canonical_path = validate_sandboxed_path(file_path)
        
        ext = canonical_path.suffix.lower()
        if ext == ".pdf":
            return StandardsParser._parse_pdf(canonical_path)
        elif ext in (".txt", ".md"):
            return StandardsParser._parse_text_or_markdown(canonical_path)
        elif ext in (".xlsx", ".xls"):
            return StandardsParser._parse_excel(canonical_path)
        else:
            raise StandardIngestError(
                f"Unsupported format '{ext}'. Standards must be one of: "
                f"{', '.join('.' + f for f in SUPPORTED_STANDARD_FORMATS)}."
            )

    @staticmethod
    def _get_color_category(rgb_val: str) -> str | None:
        """
        Classifies a cell background hexadecimal color into standard warning/success signals.
        """
        if not rgb_val or not isinstance(rgb_val, str):
            return None
            
        val = rgb_val.upper().lstrip("#")
        if len(val) == 8:
            val = val[2:] # Strip alpha channel
            
        if val in ("00000000", "FFFFFFFF", "000000", "FFFFFF", ""):
            return None
            
        try:
            r = int(val[0:2], 16)
            g = int(val[2:4], 16)
            b = int(val[4:6], 16)
            
            # Danger Red/Pink highlights (e.g. FFC7CE, F8D7DA, EF4444)
            if r > 200 and g < 155 and b < 155:
                return "RED WARNING"
            if r > 220 and g > 180 and b > 180 and r > g + 15 and r > b + 15:
                return "RED WARNING"
                
            # Success Green/Mint highlights (e.g. C6EFCE, D4EDDA, 10B981)
            if g > 170 and r < 210 and b < 210:
                return "GREEN SUCCESS"
            if g > 200 and r > 180 and b > 180 and g > r + 10 and g > b + 10:
                return "GREEN SUCCESS"
                
            # Warning Yellow/Amber highlights (e.g. FFF2CC, FFF3CD, F59E0B)
            if r > 210 and g > 190:
                if b < 170 or (b < 215 and r > b + 30 and g > b + 20):
                    return "YELLOW ATTENTION"
        except (ValueError, IndexError):
            pass
        return None

    @staticmethod
    def _read_xlsx_rows(file_path: Path, meta: dict[str, Any]) -> Iterator[_SheetRow]:
        """OOXML reader (`.xlsx`/`.xlsm`), via openpyxl."""
        import openpyxl

        # data_only so a formula cell yields its calculated value rather than "=B2*3".
        wb = openpyxl.load_workbook(file_path, data_only=True)
        meta["page_count"] = len(wb.sheetnames)

        # Embedded pictures are not read. `iter_rows()` yields cells only, so a diagram or a
        # pasted screenshot contributes no text and no chunk. Counted and reported rather than
        # silently dropped; reading them would need OCR, which is a decision, not an oversight.
        images = 0

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            images += len(getattr(sheet, "_images", ()) or ())
            for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                cells = []
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value).strip()
                    if not text:
                        continue
                    fill = None
                    if cell.fill and cell.fill.fill_type == "solid":
                        fill = getattr(cell.fill.start_color, "rgb", None)
                    cells.append(
                        _Cell(text=text, bold=bool(cell.font and cell.font.bold), fill_rgb=fill)
                    )
                if cells:
                    yield _SheetRow(sheet_name, row_idx, cells)

        meta["images_ignored"] = images

    @staticmethod
    def _read_xls_rows(file_path: Path, meta: dict[str, Any]) -> Iterator[_SheetRow]:
        """Legacy BIFF reader (`.xls`), via xlrd.

        `.xls` is not an older `.xlsx` — it is BIFF records inside an OLE2 container, sharing
        nothing with OOXML but a purpose. openpyxl reads OOXML only (its own
        `SUPPORTED_FORMATS` says so), hence a second library.

        `formatting_info=True` is what makes this reader equivalent rather than degraded: without
        it, bold and fill colour are unavailable and `.xls` would silently lose the semantic
        highlighting `.xlsx` gets — the same content producing a different corpus depending on
        which extension it was saved as. Measured on a real 30.8 MB, 18-sheet standards workbook:
        1.2 s, 67 MB peak, no different from `formatting_info=False`, so the flag is free
        here. It also reports slightly *larger* sheet extents, because styled-but-empty cells
        count toward dimensions — harmless, as empty cells are skipped below.
        """
        import xlrd

        book = xlrd.open_workbook(str(file_path), formatting_info=True)
        try:
            meta["page_count"] = book.nsheets
            # xlrd exposes no picture API at all, so unlike the xlsx reader this cannot even be
            # counted. Recorded as unknown rather than as 0, which would assert something false.
            meta["images_ignored"] = None
            meta["codepage"] = book.codepage

            colour_map = book.colour_map
            for sheet_idx in range(book.nsheets):
                sheet = book.sheet_by_index(sheet_idx)
                for row_idx in range(sheet.nrows):
                    cells = []
                    for col_idx in range(sheet.ncols):
                        value = sheet.cell_value(row_idx, col_idx)
                        if value is None:
                            continue
                        text = str(value).strip()
                        if not text:
                            continue

                        bold = False
                        fill = None
                        try:
                            xf = book.xf_list[sheet.cell_xf_index(row_idx, col_idx)]
                            font = book.font_list[xf.font_index]
                            bold = bool(font.bold) or font.weight >= 700
                            # fill_pattern 1 is "solid", mirroring openpyxl's fill_type check.
                            # For a solid fill the *pattern* colour is the visible one.
                            if xf.background.fill_pattern == 1:
                                rgb = colour_map.get(xf.background.pattern_colour_index)
                                if rgb:
                                    fill = "%02X%02X%02X" % rgb
                        except (IndexError, AttributeError, TypeError):
                            # A cell with no style record is not an error; it is an unstyled cell.
                            pass

                        cells.append(_Cell(text=text, bold=bold, fill_rgb=fill))
                    if cells:
                        yield _SheetRow(sheet.name, row_idx + 1, cells)
        finally:
            book.release_resources()

    @staticmethod
    def _parse_excel(file_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        """Chunks a workbook of either Excel format.

        The two formats differ only in how a cell is read. Everything that decides what the
        output *means* — what a fill colour signifies, that bold becomes markdown, that a row is
        joined with ` | `, that ten rows make a chunk — lives here, once, and is shared. That
        split is the whole point: a second copy of the colour rules would be free to disagree
        with the first, and the disagreement would look like a difference in the standards.
        """
        logger.info(f"Parsing Style-Aware Excel Engineering Standard: {file_path.name}")
        chunks: list[dict[str, Any]] = []
        global_metadata: dict[str, Any] = {"title": file_path.stem, "page_count": 1}

        reader = (
            StandardsParser._read_xls_rows
            if file_path.suffix.lower() == ".xls"
            else StandardsParser._read_xlsx_rows
        )

        def flush(sheet_name: str, lines: list[str], row_end: int | None) -> None:
            if not lines:
                return
            chunk_text = "\n".join(lines)
            entry_meta: dict[str, Any] = {"sheet": sheet_name, "char_count": len(chunk_text)}
            if row_end is not None:
                entry_meta["row_end"] = row_end
            chunks.append(
                {
                    "content": chunk_text,
                    "section_header": f"Sheet: {sheet_name}",
                    "metadata": entry_meta,
                }
            )

        try:
            current_sheet: str | None = None
            current_chunk: list[str] = []

            for sheet_name, row_idx, cells in reader(file_path, global_metadata):
                # Chunks never span sheets — a sheet boundary is a topic boundary here.
                if sheet_name != current_sheet:
                    flush(current_sheet or "", current_chunk, None)
                    current_chunk = []
                    current_sheet = sheet_name

                row_styles = set()
                row_data = []
                for cell in cells:
                    row_data.append(f"**{cell.text}**" if cell.bold else cell.text)
                    category = StandardsParser._get_color_category(cell.fill_rgb)
                    if category:
                        row_styles.add(category)

                content_line = " | ".join(row_data)
                if "RED WARNING" in row_styles:
                    content_line = f"[INCORRECT / DANGER FLAG] {content_line}"
                elif "GREEN SUCCESS" in row_styles:
                    content_line = f"[CORRECT / STANDARDS COMPLIANT] {content_line}"
                elif "YELLOW ATTENTION" in row_styles:
                    content_line = f"[IMPORTANT / ATTENTION REQUIRED] {content_line}"

                current_chunk.append(content_line)

                if len(current_chunk) >= 10:
                    flush(sheet_name, current_chunk, row_idx)
                    current_chunk = []

            flush(current_sheet or "", current_chunk, None)

            # Set by whichever reader ran: a count for .xlsx, None for .xls, where xlrd exposes
            # no picture API and 0 would assert something we did not check.
            images_ignored = global_metadata.get("images_ignored")
            if images_ignored:
                logger.warning(
                    f"[standards] {file_path.name}: {images_ignored} embedded image(s) were not "
                    f"read. Any rule that exists only inside a picture is absent from this "
                    f"standard's searchable content."
                )

        except StandardIngestError:
            raise
        except Exception as e:
            logger.error(f"Error parsing Excel Standard file {file_path.name}: {str(e)}")
            raise StandardIngestError(f"This workbook could not be read ({e}).")

        return chunks, global_metadata

    @staticmethod
    def _parse_pdf(file_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        logger.info(f"Parsing PDF Engineering Standard: {file_path.name}")
        chunks: list[dict[str, Any]] = []
        global_metadata: dict[str, Any] = {}

        try:
            reader = PdfReader(str(file_path))
            global_metadata["page_count"] = len(reader.pages)
            
            # Extract basic PDF info
            if reader.metadata:
                global_metadata["title"] = reader.metadata.title or file_path.stem
                global_metadata["author"] = reader.metadata.author or "Unknown"
                global_metadata["creator"] = reader.metadata.creator or "Unknown"
            else:
                global_metadata["title"] = file_path.stem

            # Extract page-by-page text
            for page_idx, page in enumerate(reader.pages):
                text = page.extract_text()
                if not text:
                    continue

                # Segment text on page by paragraphs or structural lines
                paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
                
                # If paragraph split didn't yield much, try line splits
                if len(paragraphs) <= 1:
                    paragraphs = [p.strip() for p in text.split("\n") if len(p.strip()) > 30]

                # Process each paragraph as a chunk or merge short ones
                current_chunk = ""
                section_header = None

                for p in paragraphs:
                    # Detect headers (lines that are all caps or have section numbering like 1.1, 2.0)
                    header_match = re.match(r"^([A-Z0-9\.\s]{3,40})$|^(\d+\.\d+[\s\w\-]+)$", p)
                    if header_match:
                        section_header = p
                        continue

                    if len(current_chunk) + len(p) < 800:
                        current_chunk += "\n" + p if current_chunk else p
                    else:
                        if current_chunk:
                            chunks.append({
                                "content": current_chunk.strip(),
                                "section_header": section_header,
                                "metadata": {"page_number": page_idx + 1, "char_count": len(current_chunk)}
                            })
                        current_chunk = p

                if current_chunk:
                    chunks.append({
                        "content": current_chunk.strip(),
                        "section_header": section_header,
                        "metadata": {"page_number": page_idx + 1, "char_count": len(current_chunk)}
                    })

        except StandardIngestError:
            raise
        except Exception as e:
            logger.error(f"Error parsing PDF Standard file {file_path.name}: {str(e)}")
            raise StandardIngestError(f"This PDF could not be read ({e}).")

        return chunks, global_metadata

    @staticmethod
    def _parse_text_or_markdown(file_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        logger.info(f"Parsing Text/Markdown Engineering Standard: {file_path.name}")
        chunks: list[dict[str, Any]] = []
        global_metadata: dict[str, Any] = {
            "title": file_path.stem,
            "page_count": 1
        }

        try:
            # `errors="ignore"` on a hard-coded utf-8 was silent data loss in a Japanese CAD
            # shop: a Shift-JIS standard decodes to mangled-but-non-empty text, which passes
            # every downstream emptiness check and lands a corrupted corpus nobody can see is
            # corrupted. Try the encodings this domain actually produces, strictly, and record
            # which one won so the choice is inspectable afterwards.
            content = None
            for encoding in ("utf-8", "cp932", "shift_jis", "utf-16"):
                try:
                    content = file_path.read_text(encoding=encoding)
                    global_metadata["encoding"] = encoding
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            if content is None:
                raise StandardIngestError(
                    "This file's text encoding could not be identified (tried UTF-8, CP932, "
                    "Shift-JIS, UTF-16). Re-save it as UTF-8."
                )

            # Split by headers (Markdown style # or capital sections)
            lines = content.splitlines()
            current_chunk = []
            section_header = "General"
            line_start = 1

            for line_idx, line in enumerate(lines):
                line_stripped = line.strip()
                # Markdown headers or numbered section headers
                if line_stripped.startswith("#") or re.match(r"^\d+\.\d+\s+[A-Z]", line_stripped):
                    # Save existing chunk
                    if current_chunk:
                        chunk_text = "\n".join(current_chunk).strip()
                        if chunk_text:
                            chunks.append({
                                "content": chunk_text,
                                "section_header": section_header,
                                "metadata": {"line_start": line_start, "line_end": line_idx, "char_count": len(chunk_text)}
                            })
                    current_chunk = []
                    section_header = line_stripped.lstrip("#").strip()
                    line_start = line_idx + 1
                    continue

                current_chunk.append(line)
                
                # Auto-chunking if block size gets large (approx 1000 characters)
                chunk_len = sum(len(l) for l in current_chunk)
                if chunk_len >= 1000:
                    chunk_text = "\n".join(current_chunk).strip()
                    chunks.append({
                        "content": chunk_text,
                        "section_header": section_header,
                        "metadata": {"line_start": line_start, "line_end": line_idx + 1, "char_count": len(chunk_text)}
                    })
                    current_chunk = []
                    line_start = line_idx + 2

            if current_chunk:
                chunk_text = "\n".join(current_chunk).strip()
                if chunk_text:
                    chunks.append({
                        "content": chunk_text,
                        "section_header": section_header,
                        "metadata": {"line_start": line_start, "line_end": len(lines), "char_count": len(chunk_text)}
                    })

        except StandardIngestError:
            raise
        except Exception as e:
            logger.error(f"Error parsing Text/Markdown Standard file {file_path.name}: {str(e)}")
            raise StandardIngestError(f"This text file could not be read ({e}).")

        return chunks, global_metadata
