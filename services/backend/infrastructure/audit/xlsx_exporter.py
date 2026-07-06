from pathlib import Path
from typing import Any

from ...logger import logger


class XLSXComplianceExporter:
    """
    Generates structured Excel audit sheets using openpyxl with resilient fallback mechanisms.
    """
    @staticmethod
    def generate_xlsx(output_path: Path, session: Any, drawing: Any, standard: Any, violations: list[Any]) -> Path:
        logger.info(f"Generating XLSX technical sheets to: {output_path}")
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            
            wb = openpyxl.Workbook()
            
            # Tab 1: Executive Summary
            ws_summary = wb.active
            ws_summary.title = "Executive Summary"
            ws_summary.views.sheetView[0].showGridLines = True
            
            title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            section_font = Font(name="Calibri", size=12, bold=True, color="000000")
            bold_font = Font(name="Calibri", size=11, bold=True)
            regular_font = Font(name="Calibri", size=11)
            
            navy_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            soft_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            thin_side = Side(border_style="thin", color="D9D9D9")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            # Title
            ws_summary["A1"] = "AI-2D-CHECKER: TECHNICAL AUDIT SUMMARY"
            ws_summary["A1"].font = title_font
            
            ws_summary["A3"] = "PROJECT METADATA"
            ws_summary["A3"].font = section_font
            
            metadata_rows = [
                ("Drawing Name", drawing.file_name if drawing else "Unknown"),
                ("Target Standard", standard.name if standard else "Default Standard"),
                ("Compliance Rating", f"{session.compliance_score}%"),
                ("Confidence Level", f"{session.confidence_score}%"),
                ("Total Infractions", len(violations)),
                ("Evaluation Date", session.completed_at.strftime("%Y-%m-%d %H:%M:%S") if session.completed_at else "N/A")
            ]
            
            for idx, (label, val) in enumerate(metadata_rows):
                row = 4 + idx
                ws_summary.cell(row=row, column=1, value=label).font = bold_font
                ws_summary.cell(row=row, column=1).fill = soft_blue_fill
                ws_summary.cell(row=row, column=1).border = thin_border
                
                ws_summary.cell(row=row, column=2, value=val).font = regular_font
                ws_summary.cell(row=row, column=2).border = thin_border
                
            # Tab 2: Detailed Violations
            ws_detail = wb.create_sheet(title="Audit Violations Register")
            ws_detail.views.sheetView[0].showGridLines = True
            
            headers = ["VIOLATION ID", "LAYER", "SEVERITY", "COORDINATES", "DESCRIPTION", "CONFIDENCE SCORE", "REMEDIATION STEPS"]
            for col_idx, h in enumerate(headers):
                cell = ws_detail.cell(row=1, column=col_idx+1, value=h)
                cell.font = header_font
                cell.fill = navy_fill
                cell.alignment = Alignment(horizontal="center")
                
            for idx, v in enumerate(violations):
                row = 2 + idx
                severity = getattr(v, "severity", "medium").upper()
                layer = getattr(v, "layer", "0")
                coords = f"{getattr(v, 'coordinates', [0,0])}"
                desc = getattr(v, "description", "")
                conf = f"{getattr(v, 'confidence', 1.0) * 100:.0f}%"
                remediation = f"Review design rules on layer '{layer}'. Check standard parameters."
                
                ws_detail.cell(row=row, column=1, value=f"V-{idx+1}").font = bold_font
                ws_detail.cell(row=row, column=2, value=layer).font = regular_font
                ws_detail.cell(row=row, column=3, value=severity).font = bold_font
                ws_detail.cell(row=row, column=4, value=coords).font = regular_font
                ws_detail.cell(row=row, column=5, value=desc).font = regular_font
                ws_detail.cell(row=row, column=6, value=conf).font = regular_font
                ws_detail.cell(row=row, column=7, value=remediation).font = regular_font
                
                # Apply borders
                for col_idx in range(len(headers)):
                    ws_detail.cell(row=row, column=col_idx+1).border = thin_border
                    
            # Auto-fit column widths
            for sheet in [ws_summary, ws_detail]:
                for col in sheet.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                    
            wb.save(str(output_path))
            logger.info("Successfully generated openpyxl workbook.")
            
        except ImportError:
            logger.warning("openpyxl is not installed. Generating resilient CSV technical sheet fallback.")
            fallback_text = "VIOLATION ID,LAYER,SEVERITY,DESCRIPTION,CONFIDENCE SCORE\n"
            for idx, v in enumerate(violations):
                desc_escaped = getattr(v, "description", "").replace('"', '""')
                fallback_text += f"V-{idx+1},{getattr(v, 'layer', '0')},{getattr(v, 'severity', 'medium').upper()},\"{desc_escaped}\",{getattr(v, 'confidence', 1.0) * 100:.0f}%\n"
                
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(fallback_text)
                
        return output_path
